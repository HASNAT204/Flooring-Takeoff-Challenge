"""Phase 7 — Emit the BASE BID XLSX takeoff workbook.

Mirrors the column layout used by the gold human takeoff so a reviewer can
audit side-by-side. Cost columns are intentionally left blank (with a note
in `prediction.json` explaining that costs were not estimated) — we don't
want to fake numbers we haven't computed.
"""
from __future__ import annotations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Prediction


COLUMNS = [
    ("ITEM #", 8),
    ("Drawing #", 12),
    ("DESCRIPTION", 60),
    ("QUANTITY", 12),
    ("WASTAGE %", 11),
    ("QTY WITH WASTAGE", 17),
    ("UNIT", 8),
    ("UNIT MATERIAL", 14),
    ("UNIT LABOUR", 13),
    ("UNIT EQUIPMENT", 16),
    ("TOTAL MATERIAL", 15),
    ("TOTAL LABOR", 13),
    ("TOTAL EQUIPMENT", 16),
    ("TOTAL COST", 13),
    ("CONFIDENCE", 12),
    ("METHOD", 22),
    ("SOURCE PAGE", 12),
]

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D8E1EE")
DIVISION_FILL = PatternFill("solid", fgColor="F4E5D0")
SECTION_FILL = PatternFill("solid", fgColor="EAF1FB")


def write_takeoff_xlsx(prediction: Prediction, output_path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BASE BID"

    # Header rows ------------------------------------------------------
    project_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=10)

    ws["A1"] = "FLOORING TAKEOFF — " + prediction.project
    ws["A1"].font = project_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    ws["A2"] = (f"Scale: {prediction.scale.get('ratio', 'unknown')}    "
                f"Overall confidence: {prediction.overall_confidence:.0%}    "
                f"Total flooring: {prediction.totals.get('flooring_sf', 0):.1f} SF    "
                f"Total linear: {prediction.totals.get('base_lf', 0) + prediction.totals.get('other_lf', 0):.1f} LF")
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))

    # Column widths
    for i, (_, w) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row -------------------------------------------------------
    HEADER_ROW = 4
    for i, (name, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=HEADER_ROW, column=i, value=name)
        c.font = header_font
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[HEADER_ROW].height = 30
    ws.freeze_panes = ws[f"A{HEADER_ROW + 1}"]

    # Group items by division+category --------------------------------
    by_section: dict[tuple[str, str], list] = {}
    for it in prediction.line_items:
        by_section.setdefault((it.division, it.category), []).append(it)

    row = HEADER_ROW + 1
    item_no = 1
    last_division: str | None = None
    for (division, category), items in by_section.items():
        if division != last_division:
            ws.cell(row=row, column=1, value=f"DIVISION {division}").font = Font(bold=True, size=11)
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=row, column=col).fill = DIVISION_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
            row += 1
            last_division = division
        # Section header
        ws.cell(row=row, column=1, value=category).font = Font(bold=True, italic=True)
        for col in range(1, len(COLUMNS) + 1):
            ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        row += 1
        for it in items:
            qty = it.quantity
            wpct = it.wastage_pct / 100.0
            qty_w = it.quantity_with_wastage
            ws.cell(row=row, column=1, value=item_no)
            ws.cell(row=row, column=2, value=", ".join(f"A{p}" for p in it.source_pages)
                    if it.source_pages else "")
            desc_cell = ws.cell(row=row, column=3, value=it.description)
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=4, value=qty).number_format = "#,##0.00"
            ws.cell(row=row, column=5, value=wpct).number_format = "0%"
            ws.cell(row=row, column=6, value=qty_w).number_format = "#,##0.00"
            ws.cell(row=row, column=7, value=it.unit)
            # Cost columns: blank; explicit empty cells preserve grid
            for col in (8, 9, 10, 11, 12, 13, 14):
                ws.cell(row=row, column=col, value=None)
            ws.cell(row=row, column=15, value=round(it.confidence, 2))
            ws.cell(row=row, column=16, value=it.method)
            ws.cell(row=row, column=17, value=", ".join(str(p) for p in it.source_pages))
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=row, column=col).border = BORDER
                if col == 3:
                    continue
                ws.cell(row=row, column=col).alignment = Alignment(vertical="top", horizontal="left" if col >= 14 else "center")
            ws.row_dimensions[row].height = max(48, 16 + 16 * (it.description.count("\n") + 1))
            row += 1
            item_no += 1

    # Totals row -------------------------------------------------------
    row += 1
    ws.cell(row=row, column=1, value="TOTALS").font = Font(bold=True)
    ws.cell(row=row, column=4, value=sum(it.quantity for it in prediction.line_items if it.unit == "SF")).number_format = "#,##0.00"
    ws.cell(row=row, column=7, value="SF")
    row += 1
    ws.cell(row=row, column=4, value=sum(it.quantity for it in prediction.line_items if it.unit == "LF")).number_format = "#,##0.00"
    ws.cell(row=row, column=7, value="LF")

    # Notes row --------------------------------------------------------
    row += 2
    ws.cell(row=row, column=1, value=(
        "Costs (material/labour/equipment) intentionally blank — this product computes "
        "quantities only. Confidence column is the area-weighted confidence of the underlying "
        "measurement, computed by the pipeline."
    )).font = Font(italic=True, color="666666", size=9)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    ws.row_dimensions[row].height = 32

    wb.save(str(output_path))
