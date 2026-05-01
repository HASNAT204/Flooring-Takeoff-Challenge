"""Loads the gold takeoff XLSX. THIS IS THE ONLY MODULE THAT READS gold/.

Keeping the file path open here, and grepping `app/` for the substring
``gold`` (see tests/test_gold_leakage.py) is how we enforce the rule that
the prediction pipeline never sees the gold output.
"""
from __future__ import annotations
import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


CODE_PATTERNS = {
    "V-1": [r"\bV-?1\b", r"VINYL COMPOSITION", r"ARMSTRONG.*EXCELON", r"IMPERIAL TEXTURE"],
    "T-1": [r"\bT-?1\b", r"LIGHT COLOR.*PORCELAIN", r"WHITE OAK 00100", r"PATCRAFT.*CS50H", r'8".*48".*PORCELAIN'],
    "T-2": [r"\bT-?2\b", r"DARK COLOR.*PORCELAIN", r"WALNUT 0700", r"CREEKWOOD"],
    "W-1": [r"\bW-?1\b", r"LOVESAC OAK", r"HARBOR"],
    "W-2": [r"\bW-?2\b", r"LOVESAC HICKORY", r"TRANQUILITY"],
    "RB":  [r"RUBBER BASE"],
    "MDF-BASE": [r"MDF.*BASE", r"PAINTED WOOD BASE", r"PAINTED.*MDF"],
    "EDGE-TRIM": [r"EDGE TRIM", r"SCHLUTER"],
    "TRANSITION-VT": [r"TRANSITION", r"VINYL.*PORCELAIN"],
}


def _classify_description(desc: str) -> Optional[str]:
    if not desc:
        return None
    desc_u = desc.upper()
    # Order matters — most specific first
    for code, patterns in CODE_PATTERNS.items():
        for p in patterns:
            if re.search(p, desc_u):
                return code
    return None


def load_gold(xlsx_path: str | Path) -> dict[str, dict]:
    """Return a dict keyed by estimator code with gold quantities.

    Each value is ``{quantity, wastage_pct, qty_with_wastage, unit, description}``.
    Items that can't be classified are returned under ``_unclassified`` key.
    """
    wb = load_workbook(str(xlsx_path), data_only=True)
    out: dict[str, dict] = {}
    unclassified: list[dict] = []
    for ws in wb.worksheets:
        # find header row (first row containing both "QUANTITY" and "UNIT")
        header_row = None
        for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_text = " ".join(str(c) if c else "" for c in row).upper()
            if "QUANTITY" in row_text and "UNIT" in row_text:
                header_row = ri
                header = [str(c).upper().strip() if c else "" for c in row]
                break
        if header_row is None:
            continue

        # Map column names to indexes
        def col(*names: str) -> Optional[int]:
            for i, h in enumerate(header):
                for n in names:
                    if n in h:
                        return i
            return None

        c_desc = col("DESCRIPTION")
        c_qty = col("QUANTITY")
        c_unit = col("UNIT")
        c_waste = col("WASTAGE")
        c_qty_w = col("QTY WITH WASTAGE", "QUANTITY WITH WASTAGE")

        if c_desc is None or c_qty is None or c_unit is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            desc = row[c_desc]
            qty = row[c_qty]
            unit = row[c_unit]
            if desc is None or qty is None or unit is None:
                continue
            try:
                qty = float(qty)
            except (TypeError, ValueError):
                continue
            unit = str(unit).strip().upper()
            if unit not in ("SF", "LF", "FT", "EA", "LS"):
                continue
            if unit == "FT":
                unit = "LF"
            desc_str = str(desc)
            code = _classify_description(desc_str)
            wastage = row[c_waste] if c_waste is not None else 0
            qty_w = row[c_qty_w] if c_qty_w is not None else qty
            entry = {
                "description": desc_str,
                "quantity": round(qty, 3),
                "wastage_pct": float(wastage) if wastage else 0.0,
                "quantity_with_wastage": float(qty_w) if qty_w else qty,
                "unit": unit,
            }
            if code is None:
                unclassified.append(entry)
                continue
            # If the same code appears more than once (e.g. management fee) sum
            if code in out:
                out[code]["quantity"] = round(out[code]["quantity"] + qty, 3)
                out[code]["quantity_with_wastage"] += entry["quantity_with_wastage"]
            else:
                out[code] = entry
    if unclassified:
        out["_unclassified"] = unclassified  # type: ignore[assignment]
    return out
